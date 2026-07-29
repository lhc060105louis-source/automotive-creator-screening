from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models import Shortlist
from app.assessment import commercial_grade
from app.config import COMMERCIAL_WEIGHTS, RISK_WEIGHTS
from app.models import Kol

HEADERS = [
    "Priority",
    "KOL",
    "Country",
    "Platform",
    "Handle",
    "Followers",
    "Commercial Score",
    "Commercial Completeness",
    "Risk Score",
    "Risk Level",
    "Risk Completeness",
    "Recommendation",
]


def export_shortlist(shortlist: Shortlist) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KOL Candidates"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")

    for item in sorted(
        shortlist.items,
        key=lambda value: (value.priority is None, value.priority or 0, value.id),
    ):
        kol = item.kol
        summary = kol.score_summary
        sheet.append(
            [
                item.priority,
                kol.name or kol.handle,
                kol.country,
                kol.platform,
                kol.handle,
                kol.followers,
                summary.commercial_score if summary else None,
                summary.commercial_completeness if summary else 0,
                summary.risk_score if summary else None,
                summary.risk_level if summary else None,
                summary.risk_completeness if summary else 0,
                item.recommendation,
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {
        "A": 10,
        "B": 24,
        "C": 10,
        "D": 14,
        "E": 20,
        "F": 14,
        "G": 20,
        "H": 24,
        "I": 14,
        "J": 14,
        "K": 20,
        "L": 36,
    }.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 8).number_format = "0%"
        sheet.cell(row, 11).number_format = "0%"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


KOL_EXPORT_HEADERS = [
    "KOL", "Country", "Platform", "Handle", "Account ID", "Profile URL",
    "Language", "Content Categories", "Followers", "Engagement Rate",
    "Audience Country Ratio",
    *[name.replace("_", " ").title() for name in COMMERCIAL_WEIGHTS],
    "Commercial Score", "Commercial Grade", "Commercial Completeness",
    *[name.replace("_", " ").title() for name in RISK_WEIGHTS],
    "Risk Score", "Risk Level", "Risk Completeness", "Source", "Updated Time",
]


def export_kols(kols: list[Kol]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KOLs"
    sheet.append(KOL_EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")

    for kol in kols:
        summary = kol.score_summary
        scores = {
            (record.score_type, record.dimension): record.final_score
            for record in kol.score_records
        }
        source = kol.assessment_input.source if kol.assessment_input else None
        sheet.append([
            kol.name or kol.handle, kol.country, kol.platform, kol.handle,
            kol.platform_account_id, kol.profile_url, kol.language,
            kol.content_categories, kol.followers, kol.average_engagement_rate,
            kol.audience_country_ratio,
            *[scores.get(("commercial", dimension)) for dimension in COMMERCIAL_WEIGHTS],
            summary.commercial_score if summary else None,
            commercial_grade(summary.commercial_score) if summary else None,
            summary.commercial_completeness if summary else 0,
            *[scores.get(("risk", dimension)) for dimension in RISK_WEIGHTS],
            summary.risk_score if summary else None,
            summary.risk_level if summary else None,
            summary.risk_completeness if summary else 0,
            source, kol.updated_at,
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2, 36
        )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
