import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import COMMERCIAL_WEIGHTS, RISK_WEIGHTS
from app.models import (
    AssessmentInput,
    ImportError,
    ImportJob,
    Kol,
    KolScoreSummary,
    ScoreRecord,
)
from app.assessment import calculate_assessment
from app.identity import find_existing_kol
from app.market import normalize_market
from app.services.scoring import risk_level, summarize_dimensions

HEADER_ALIASES = {
    "平台": "platform",
    "国家": "country",
    "账号": "handle",
    "账号id": "platform_account_id",
    "名称": "name",
    "主页": "profile_url",
    "语言": "language",
    "内容类别": "content_categories",
    "粉丝量": "followers",
    "互动率": "average_engagement_rate",
    "受众国家占比": "audience_country_ratio",
    "受众匹配度": "audience_fit",
    "内容相关性与专业度": "content_relevance",
    "互动质量": "interaction_quality",
    "voc反馈价值": "voc_value",
    "商业效率": "commercial_efficiency",
    "品牌适配度": "brand_fit",
    "合作可执行性": "execution_capability",
    "历史负面舆情": "historical_controversy",
    "广告披露合规风险": "ad_disclosure",
    "竞品冲突": "competitor_conflict",
    "虚假流量风险": "fake_traffic",
    "数据与隐私风险": "data_privacy",
    "未成年人/敏感受众风险": "sensitive_audience",
    "可持续/技术声明风险": "sustainability_claims",
    "合作执行风险": "execution_risk",
}

COMPLETE_COMMERCIAL_HEADERS = {
    "KOL名称": "name",
    "网红名称": "name",
    "主要市场(DE/GB/FR/MULTI)": "country",
    "内容方向(review/ev/luxury/family/tech/lifestyle)": "contentDirection",
    "目标品牌类型(premium/mainstream/ev-brand)": "targetBrand",
    "目标市场受众占比%": "geo", "语言能力(1/2/3)": "lang",
    "汽车兴趣受众%": "autoInterest", "25-55岁受众%": "age",
    "收入水平(high/mid/low)": "income", "汽车内容专注%": "focus",
    "测评深度(deep/mid/surface)": "depth", "专业可信度(high/mid/low)": "credibility",
    "ERR%（YouTube API自动抓取）": "err", "完播率%": "completion",
    "评论质量(high/mid/low)": "commentQuality", "分享收藏比%": "shareSave",
    "VOC话题深度(high/mid/low)": "vocDepth", "VOC负面识别(high/mid/low)": "vocNeg",
    "历史车主反馈(yes/sometimes/no)": "vocHistory", "CPM报价€": "cpm",
    "行业基准CPM€": "benchCpm", "内容复用权(full/limited/none)": "reuse",
    "排他要求(none/soft/hard)": "exclusive", "品牌调性匹配(match/neutral/conflict)": "brandTone",
    "历史合作调性(match/neutral/conflict)": "histTone", "内容风格一致性(high/mid/low)": "styleConsist",
    "履约率%": "fulfill", "Brief配合度(high/mid/low)": "briefCoop",
    "数据复盘意愿(active/passive/refuse)": "dataReady",
}

COMPLETE_RISK_HEADERS = {
    "KOL名称": "name", "主要市场(DE/GB/FR/MULTI)": "country",
    "重大负面事件(none/minor/serious/critical)": "incident", "虚假宣传记录(none/minor/serious)": "falsead",
    "舆情传播范围(none/local/wide)": "sentiment", "广告标注习惯(always/sometimes/never)": "adlabel",
    "平台监管处罚(none/warning/penalty)": "penalty", "合规意愿(high/mid/low)": "compliance",
    "竞品绑定状态(none/nonexclusive/exclusive/ambassador)": "competitor", "近期竞品内容%": "compcontentpct",
    "竞品品牌级别(none/indirect/direct)": "complevel", "僵尸粉比例%": "fakepct",
    "粉丝暴涨记录(none/once/multiple)": "spikegrowth", "评论模板化(normal/some/heavy)": "templatecomment",
    "GDPR违规记录(none/minor/serious)": "gdpr", "数据使用规范(compliant/unclear/violation)": "datause",
    "未成年受众%": "minorpct", "内容适龄性(suitable/partial/unsuitable)": "agesuit",
    "历史夸大声明(none/minor/serious)": "exaggerate", "自动驾驶声明风险(none/cautious/exaggerated)": "adas",
    "技术准确性(high/mid/low)": "techaccuracy", "历史延期删帖(none/occasional/frequent)": "latedelete",
    "Brief拒绝修改(cooperative/friction/refuse)": "briefreject",
}

BASE_FIELDS = {
    "name",
    "platform",
    "platform_account_id",
    "handle",
    "profile_url",
    "country",
    "language",
    "content_categories",
    "followers",
    "average_engagement_rate",
    "audience_country_ratio",
}


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    lowered = text.lower()
    return HEADER_ALIASES.get(lowered, HEADER_ALIASES.get(text, lowered))


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
    if not reader.fieldnames:
        raise ValueError("file has no header row")
    return [
        {normalize_header(key): value for key, value in row.items() if key is not None}
        for row in reader
    ]


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(value) for value in next(values)]
    except StopIteration as exc:
        raise ValueError("file has no header row") from exc
    return [dict(zip(headers, row, strict=False)) for row in values]


def _normalized_sheet_name(value: str) -> str:
    return "".join(value.split()).lower()


def _complete_rows(content: bytes) -> list[dict[str, Any]] | None:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheets = {_normalized_sheet_name(sheet.title): sheet for sheet in workbook.worksheets}
    commercial = sheets.get("商业价值模型")
    risk = sheets.get("风险评估模型")
    if commercial is None or risk is None:
        return None

    def mapped_rows(sheet, mapping):
        values = sheet.iter_rows(values_only=True)
        headers = next(values, ())
        keys = [mapping.get(str(header or "").strip()) for header in headers]
        return [
            {key: value for key, value in zip(keys, row, strict=False) if key}
            for row in values if any(value is not None for value in row)
        ]

    commercial_rows = mapped_rows(commercial, COMPLETE_COMMERCIAL_HEADERS)
    risk_by_name = {
        str(row.get("name") or "").strip().casefold(): row
        for row in mapped_rows(risk, COMPLETE_RISK_HEADERS) if row.get("name")
    }
    rows = []
    for row in commercial_rows:
        name = _clean_text(row.get("name"))
        joined = dict(row)
        joined["_risk_inputs"] = risk_by_name.get((name or "").casefold(), {})
        joined["_complete_package"] = True
        rows.append(joined)
    return rows


def read_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return _rows_from_csv(content)
    if extension == ".xlsx":
        return _complete_rows(content) or _rows_from_xlsx(content)
    raise ValueError("unsupported file type")


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_number(value: object, *, integer: bool = False) -> float | int | None:
    text = _clean_text(value)
    if text is None:
        return None
    number = float(text.replace(",", ""))
    return int(number) if integer else number


def _clean_ratio(value: object) -> float | None:
    text = _clean_text(value)
    if text is None:
        return None
    if text.endswith("%"):
        return float(text[:-1].replace(",", "")) / 100
    number = float(text.replace(",", ""))
    return number / 100 if number > 1 else number


def _clean_score(value: object) -> float | None:
    score = _clean_number(value)
    if score is None:
        return None
    if not 0 <= score <= 100:
        raise ValueError("score must be between 0 and 100")
    return float(score)


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    if row.get("_complete_package"):
        name = _clean_text(row.get("name"))
        country = _clean_text(row.get("country"))
        if not name:
            raise ValueError("KOL name is required")
        if not country:
            raise ValueError("country is required")
        if country.upper() not in {"DE", "GB", "FR", "MULTI"}:
            raise ValueError("country must be DE, GB, FR, or MULTI")
        commercial = {
            key: (None if str(value or "").strip() == "← YouTube API" else value)
            for key, value in row.items()
            if key not in {"name", "country", "_risk_inputs", "_complete_package"}
        }
        commercial["contractFlex"] = None
        risk = {
            key: value for key, value in row["_risk_inputs"].items()
            if key not in {"name", "country"}
        }
        return {
            "platform": "YouTube", "country": country.upper(), "handle": name,
            "platform_account_id": None, "name": name, "profile_url": None,
            "language": None, "content_categories": _clean_text(row.get("contentDirection")),
            "followers": None, "average_engagement_rate": None,
            "audience_country_ratio": _clean_ratio(row.get("geo")),
            "_commercial_inputs": commercial, "_risk_inputs": risk,
        }
    platform = _clean_text(row.get("platform"))
    country_text = _clean_text(row.get("country"))
    handle = _clean_text(row.get("handle"))
    account_id = _clean_text(row.get("platform_account_id"))
    if not platform:
        raise ValueError("platform is required")
    if not country_text:
        raise ValueError("country is required")
    country = normalize_market(country_text)
    if not handle and not account_id:
        raise ValueError("handle or platform_account_id is required")

    cleaned = {
        "platform": platform,
        "country": country,
        "handle": handle,
        "platform_account_id": account_id,
        "name": _clean_text(row.get("name")),
        "profile_url": _clean_text(row.get("profile_url")),
        "language": _clean_text(row.get("language")),
        "content_categories": _clean_text(row.get("content_categories")),
        "followers": _clean_number(row.get("followers"), integer=True),
        "average_engagement_rate": _clean_ratio(row.get("average_engagement_rate")),
        "audience_country_ratio": _clean_ratio(row.get("audience_country_ratio")),
    }
    for dimension in COMMERCIAL_WEIGHTS | RISK_WEIGHTS:
        cleaned[dimension] = _clean_score(row.get(dimension))
    return cleaned


def _find_existing(session: Session, row: dict[str, Any]) -> Kol | None:
    return find_existing_kol(
        session, platform=row["platform"], platform_account_id=row.get("platform_account_id"),
        profile_url=row.get("profile_url"), handle=row.get("handle"),
    )


def _upsert_scores(
    session: Session,
    kol: Kol,
    row: dict[str, Any],
    source: str,
) -> None:
    for score_type, weights in (
        ("commercial", COMMERCIAL_WEIGHTS),
        ("risk", RISK_WEIGHTS),
    ):
        for dimension in weights:
            score = row.get(dimension)
            if score is None:
                continue
            record = session.scalar(
                select(ScoreRecord).where(
                    ScoreRecord.kol_id == kol.id,
                    ScoreRecord.score_type == score_type,
                    ScoreRecord.dimension == dimension,
                )
            )
            if record is None:
                record = ScoreRecord(
                    kol_id=kol.id,
                    score_type=score_type,
                    dimension=dimension,
                )
                session.add(record)
            record.auto_score = score
            record.source = source


def refresh_summary(session: Session, kol: Kol) -> KolScoreSummary:
    records = list(
        session.scalars(select(ScoreRecord).where(ScoreRecord.kol_id == kol.id))
    )
    grouped: dict[str, dict[str, object]] = {"commercial": {}, "risk": {}}
    for record in records:
        grouped[record.score_type][record.dimension] = {
            "auto": record.auto_score,
            "manual": record.manual_score,
        }

    commercial = summarize_dimensions(grouped["commercial"], COMMERCIAL_WEIGHTS)
    risk = summarize_dimensions(grouped["risk"], RISK_WEIGHTS)
    summary = session.scalar(
        select(KolScoreSummary).where(KolScoreSummary.kol_id == kol.id)
    )
    if summary is None:
        summary = KolScoreSummary(kol_id=kol.id)
        session.add(summary)
    summary.commercial_score = commercial.score
    summary.commercial_completeness = commercial.completeness
    summary.commercial_status = commercial.status
    summary.risk_score = risk.score
    summary.risk_completeness = risk.completeness
    summary.risk_status = risk.status
    summary.risk_level = risk_level(risk.score)
    return summary


def persist_assessment(
    session: Session,
    kol: Kol,
    commercial_inputs: dict[str, Any],
    risk_inputs: dict[str, Any],
    *,
    source: str | None,
) -> AssessmentInput:
    """Persist raw inputs and refresh every calculable automatic dimension."""
    result = calculate_assessment(commercial_inputs, risk_inputs)
    raw = session.get(AssessmentInput, kol.id)
    if raw is None:
        raw = AssessmentInput(kol_id=kol.id)
        session.add(raw)
    raw.commercial_inputs = dict(commercial_inputs)
    raw.risk_inputs = dict(risk_inputs)
    raw.flags = list(result.flags)
    if source is not None:
        raw.source = source

    for score_type, dimensions in (
        ("commercial", result.commercial_dimensions),
        ("risk", result.risk_dimensions),
    ):
        for dimension, calculated in dimensions.items():
            record = session.scalar(
                select(ScoreRecord).where(
                    ScoreRecord.kol_id == kol.id,
                    ScoreRecord.score_type == score_type,
                    ScoreRecord.dimension == dimension,
                )
            )
            if record is None:
                record = ScoreRecord(
                    kol_id=kol.id,
                    score_type=score_type,
                    dimension=dimension,
                )
                session.add(record)
            record.auto_score = calculated.score
            if source is not None:
                record.source = source
                record.evidence = "\n".join(calculated.evidence) or None
    session.flush()
    # Summary is derived from persisted records so manual scores override automatic
    # scores and unavailable dimensions remain excluded from the denominator.
    refresh_summary(session, kol)
    return raw


def import_file(
    session: Session,
    filename: str,
    content: bytes,
) -> ImportJob:
    rows = read_rows(filename, content)
    safe_filename = Path(filename.replace("\\", "/")).name
    job = ImportJob(filename=safe_filename, status="processing", total_rows=len(rows))
    session.add(job)
    session.flush()

    for row_number, raw_row in enumerate(rows, start=2):
        try:
            created = False
            with session.begin_nested():
                row = _normalize_row(raw_row)
                kol = _find_existing(session, row)
                if kol is None:
                    kol = Kol(**{key: row[key] for key in BASE_FIELDS})
                    session.add(kol)
                    session.flush()
                    created = True
                else:
                    for key in BASE_FIELDS:
                        if row[key] is not None:
                            setattr(kol, key, row[key])
                _upsert_scores(session, kol, row, f"import:{safe_filename}")
                if "_commercial_inputs" in row:
                    persist_assessment(
                        session, kol, row["_commercial_inputs"], row["_risk_inputs"],
                        source=f"import:{safe_filename}",
                    )
                session.flush()
                refresh_summary(session, kol)
            if created:
                job.created_count += 1
            else:
                job.updated_count += 1
        except (TypeError, ValueError) as exc:
            job.failed_count += 1
            job.errors.append(
                ImportError(row_number=row_number, message=str(exc))
            )

    job.status = "completed"
    session.commit()
    session.refresh(job)
    return job
