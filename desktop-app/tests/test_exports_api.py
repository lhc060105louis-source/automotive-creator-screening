from io import BytesIO

from openpyxl import load_workbook

from test_importer import complete_package_workbook_bytes, upload_csv


def test_export_respects_same_country_filter_as_kol_search(client):
    upload_csv(client, "platform,country,handle,name,followers\nYouTube,FR,@fr,French EV,10\nYouTube,DE,@de,German EV,20\n")

    response = client.get("/exports/kols.xlsx", params={"country": "FR"})

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    rows = list(workbook.active.values)
    assert len(rows) == 2
    assert rows[1][rows[0].index("Country")] == "FR"
    assert "Token" not in rows[0]
    assert "Local Path" not in rows[0]


def test_export_supports_all_kol_search_filters(client):
    upload_csv(client, "platform,country,handle,name,followers,language,audience_fit,historical_controversy\nYouTube,FR,@one,Target EV,1000,fr,90,10\nTikTok,FR,@two,Other,100,fr,40,70\n")

    response = client.get(
        "/exports/kols.xlsx",
        params={"keyword": "Target", "country": "FR", "platform": "youtube", "language": "fr", "min_followers": 500, "max_followers": 2000, "min_commercial_score": 80, "max_risk_score": 20, "risk_level": "low", "min_completeness": .1},
    )

    assert response.status_code == 200
    rows = list(load_workbook(BytesIO(response.content), read_only=True).active.values)
    assert len(rows) == 2
    assert rows[1][rows[0].index("KOL")] == "Target EV"
    assert {"Audience Fit", "Execution Capability", "Historical Controversy", "Execution Risk", "Commercial Grade", "Source", "Updated Time"}.issubset(rows[0])


def test_export_source_never_contains_uploaded_local_path(client):
    filename = "BYD_Xpeng_KOL评估数据.xlsx"
    uploaded_name = f"/Users/private/secret/{filename}"
    response = client.post(
        "/imports", files={"file": (uploaded_name, complete_package_workbook_bytes())}
    )
    assert response.status_code == 201

    exported = client.get("/exports/kols.xlsx", params={"keyword": "Max Torques"})
    rows = list(load_workbook(BytesIO(exported.content), read_only=True).active.values)
    source = rows[1][rows[0].index("Source")]
    assert source == f"import:{filename}"
    assert "/Users/" not in source
