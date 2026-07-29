from io import BytesIO

from openpyxl import Workbook


COMPLETE_COMMERCIAL_HEADERS = [
    "KOL名称", "主要市场(DE/GB/FR/MULTI)",
    "内容方向(review/ev/luxury/family/tech/lifestyle)",
    "目标品牌类型(premium/mainstream/ev-brand)", "目标市场受众占比%",
    "语言能力(1/2/3)", "汽车兴趣受众%", "25-55岁受众%",
    "收入水平(high/mid/low)", "汽车内容专注%", "测评深度(deep/mid/surface)",
    "专业可信度(high/mid/low)", "ERR%（YouTube API自动抓取）", "完播率%",
    "评论质量(high/mid/low)", "分享收藏比%", "VOC话题深度(high/mid/low)",
    "VOC负面识别(high/mid/low)", "历史车主反馈(yes/sometimes/no)", "CPM报价€",
    "行业基准CPM€", "内容复用权(full/limited/none)", "排他要求(none/soft/hard)",
    "品牌调性匹配(match/neutral/conflict)", "历史合作调性(match/neutral/conflict)",
    "内容风格一致性(high/mid/low)", "履约率%", "Brief配合度(high/mid/low)",
    "数据复盘意愿(active/passive/refuse)",
]

COMPLETE_RISK_HEADERS = [
    "KOL名称", "主要市场(DE/GB/FR/MULTI)",
    "重大负面事件(none/minor/serious/critical)", "虚假宣传记录(none/minor/serious)",
    "舆情传播范围(none/local/wide)", "广告标注习惯(always/sometimes/never)",
    "平台监管处罚(none/warning/penalty)", "合规意愿(high/mid/low)",
    "竞品绑定状态(none/nonexclusive/exclusive/ambassador)", "近期竞品内容%",
    "竞品品牌级别(none/indirect/direct)", "僵尸粉比例%",
    "粉丝暴涨记录(none/once/multiple)", "评论模板化(normal/some/heavy)",
    "GDPR违规记录(none/minor/serious)", "数据使用规范(compliant/unclear/violation)",
    "未成年受众%", "内容适龄性(suitable/partial/unsuitable)",
    "历史夸大声明(none/minor/serious)", "自动驾驶声明风险(none/cautious/exaggerated)",
    "技术准确性(high/mid/low)", "历史延期删帖(none/occasional/frequent)",
    "Brief拒绝修改(cooperative/friction/refuse)",
]


def complete_package_workbook_bytes() -> bytes:
    workbook = Workbook()
    commercial = workbook.active
    commercial.title = "商业价值模型"
    commercial.append(COMPLETE_COMMERCIAL_HEADERS)
    commercial_values = {
        "KOL名称": "Max Torques", "主要市场(DE/GB/FR/MULTI)": "GB",
        "内容方向(review/ev/luxury/family/tech/lifestyle)": "ev",
        "目标市场受众占比%": 78, "语言能力(1/2/3)": 3,
        "汽车兴趣受众%": 85, "25-55岁受众%": 72,
        "收入水平(high/mid/low)": "high", "ERR%（YouTube API自动抓取）": "← YouTube API",
    }
    commercial.append([commercial_values.get(header) for header in COMPLETE_COMMERCIAL_HEADERS])

    risk = workbook.create_sheet("风险评估模型")
    risk.append(COMPLETE_RISK_HEADERS)
    risk_values = {
        "KOL名称": "Max Torques", "主要市场(DE/GB/FR/MULTI)": "GB",
        "重大负面事件(none/minor/serious/critical)": "none",
        "历史延期删帖(none/occasional/frequent)": "none",
        "Brief拒绝修改(cooperative/friction/refuse)": "cooperative",
    }
    risk.append([risk_values.get(header) for header in COMPLETE_RISK_HEADERS])

    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def upload_csv(client, text: str, filename: str = "kols.csv"):
    return client.post(
        "/imports",
        files={"file": (filename, text.encode("utf-8"), "text/csv")},
    )


def test_import_csv_creates_kol_and_scores(client):
    csv_data = (
        "platform,country,handle,name,followers,audience_fit,"
        "historical_controversy\n"
        "YouTube,UK,@alex,Alex EV,120000,85,10\n"
    )

    response = upload_csv(client, csv_data)

    assert response.status_code == 201
    assert response.json()["created"] == 1
    detail = client.get("/kols/1").json()
    assert detail["handle"] == "@alex"
    assert detail["score_summary"]["commercial_score"] == 85
    assert detail["score_summary"]["commercial_completeness"] == 0.2
    assert detail["score_summary"]["risk_score"] == 10


def test_import_updates_duplicate_handle(client):
    first = "platform,country,handle,followers\nYouTube,UK,@alex,100\n"
    second = "platform,country,handle,followers\nYouTube,UK,@alex,200\n"
    upload_csv(client, first, "first.csv")

    response = upload_csv(client, second, "second.csv")

    assert response.json()["updated"] == 1
    kols = client.get("/kols").json()
    assert len(kols) == 1
    assert kols[0]["followers"] == 200


def test_bad_row_does_not_block_valid_row(client):
    csv_data = (
        "platform,country,handle\n"
        "YouTube,UK,@valid\n"
        "TikTok,,@invalid\n"
    )

    response = upload_csv(client, csv_data)

    body = response.json()
    assert body["created"] == 1
    assert body["failed"] == 1
    job = client.get(f"/imports/{body['job_id']}")
    assert job.status_code == 200
    assert job.json()["errors"][0]["row_number"] == 3
    assert "country is required" in job.json()["errors"][0]["message"]


def test_chinese_headers_and_germany_are_normalized(client):
    csv_data = "平台,国家,账号,粉丝量\nYouTube,德国,@auto,10000\n"

    response = upload_csv(client, csv_data)

    assert response.status_code == 201
    kol = client.get("/kols").json()[0]
    assert kol["country"] == "DE"


def test_import_xlsx(client):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["platform", "country", "handle", "followers"])
    sheet.append(["TikTok", "英国", "@evuk", 5000])
    content = BytesIO()
    workbook.save(content)

    response = client.post(
        "/imports",
        files={
            "file": (
                "kols.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    assert response.json()["created"] == 1


def test_unsupported_extension_returns_415(client):
    response = client.post(
        "/imports",
        files={"file": ("kols.txt", b"x", "text/plain")},
    )

    assert response.status_code == 415


def test_imports_complete_package_workbook_and_maps_raw_assessment_inputs(client):
    response = client.post(
        "/imports",
        files={"file": ("BYD_Xpeng_KOL评估数据.xlsx", complete_package_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    assert response.json() == {
        "job_id": 1, "total_rows": 1, "created": 1, "updated": 0, "failed": 0
    }
    match = client.get("/kols", params={"keyword": "Max Torques"}).json()[0]
    detail = client.get(f"/kols/{match['id']}").json()
    assert detail["country"] == "GB"
    assert detail["content_categories"] == "ev"
    assert detail["average_engagement_rate"] is None
    assert detail["commercial_inputs"]["err"] is None
    assert detail["commercial_inputs"]["geo"] == 78
    assert detail["commercial_inputs"]["contractFlex"] is None
    assert len(detail["commercial_inputs"]) == 28
    assert detail["risk_inputs"]["incident"] == "none"
    assert detail["risk_inputs"]["briefreject"] == "cooperative"
    assert len(detail["risk_inputs"]) == 21
    assert detail["score_summary"]["commercial_score"] is not None
    assert detail["score_summary"]["risk_score"] is not None


def test_complete_package_bad_row_does_not_block_valid_rows(client):
    workbook = Workbook()
    commercial = workbook.active
    commercial.title = " 商业价值模型 "
    commercial.append(["网红名称", "主要市场(DE/GB/FR/MULTI)", "目标市场受众占比%"])
    commercial.append(["Valid", "GB", 70])
    commercial.append(["Broken", None, 50])
    risk = workbook.create_sheet("风险 评估模型")
    risk.append(["KOL名称", "主要市场(DE/GB/FR/MULTI)", "重大负面事件(none/minor/serious/critical)"])
    risk.append(["Valid", "GB", "none"])
    content = BytesIO()
    workbook.save(content)

    response = client.post("/imports", files={"file": ("complete.xlsx", content.getvalue())})

    assert response.status_code == 201
    assert response.json()["created"] == 1
    assert response.json()["failed"] == 1


def test_assessment_failure_rolls_back_whole_row_and_next_row_succeeds(client):
    workbook = Workbook()
    commercial = workbook.active
    commercial.title = "商业价值模型"
    commercial.append([
        "KOL名称", "主要市场(DE/GB/FR/MULTI)", "CPM报价€", "行业基准CPM€",
        "内容复用权(full/limited/none)", "排他要求(none/soft/hard)",
    ])
    commercial.append(["Atomic Failure", "GB", 10, 0, "full", "none"])
    commercial.append(["Still Valid", "GB", None, None, None, None])
    risk = workbook.create_sheet("风险评估模型")
    risk.append(["KOL名称", "主要市场(DE/GB/FR/MULTI)", "僵尸粉比例%"])
    risk.append(["Atomic Failure", "GB", "not-a-number"])
    risk.append(["Still Valid", "GB", None])
    content = BytesIO()
    workbook.save(content)

    response = client.post("/imports", files={"file": ("atomic.xlsx", content.getvalue())})

    assert response.status_code == 201
    assert response.json()["created"] == 1
    assert response.json()["failed"] == 1
    assert [kol["name"] for kol in client.get("/kols").json()] == ["Still Valid"]


def test_assessment_failure_rolls_back_existing_kol_update(client):
    upload_csv(
        client,
        "platform,country,handle,name,followers\nYouTube,UK,Existing,Existing,123\n",
    )
    workbook = Workbook()
    commercial = workbook.active
    commercial.title = "商业价值模型"
    commercial.append([
        "KOL名称", "主要市场(DE/GB/FR/MULTI)", "内容方向(review/ev/luxury/family/tech/lifestyle)",
    ])
    commercial.append(["Existing", "GB", "ev"])
    risk = workbook.create_sheet("风险评估模型")
    risk.append(["KOL名称", "主要市场(DE/GB/FR/MULTI)", "僵尸粉比例%"])
    risk.append(["Existing", "GB", "not-a-number"])
    content = BytesIO()
    workbook.save(content)

    response = client.post("/imports", files={"file": ("atomic-update.xlsx", content.getvalue())})

    assert response.json()["updated"] == 0
    assert response.json()["failed"] == 1
    existing = client.get("/kols", params={"keyword": "Existing"}).json()[0]
    assert existing["country"] == "GB"
    assert existing["content_categories"] is None
    assert existing["followers"] == 123
def test_import_deduplicates_canonical_youtube_url_and_handle_case(client):
    first = "platform,country,handle,profile_url,followers\nYouTube,GB,@Case,https://www.youtube.com/@Case/,100\n"
    second = "platform,country,handle,profile_url,followers\nyoutube,UK,@other,http://youtube.com/@case?x=1,200\n"
    assert upload_csv(client, first).json()["created"] == 1
    assert upload_csv(client, second).json()["updated"] == 1
    assert len(client.get("/kols").json()) == 1
