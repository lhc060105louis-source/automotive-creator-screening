from io import BytesIO

from openpyxl import load_workbook


def seed_two_kols(client) -> list[int]:
    csv_data = (
        "platform,country,handle,name,audience_fit,historical_controversy\n"
        "YouTube,UK,@alex,Alex EV,85,10\n"
        "TikTok,DE,@mobil,Mobil DE,75,30\n"
    )
    response = client.post(
        "/imports",
        files={"file": ("seed.csv", csv_data.encode(), "text/csv")},
    )
    assert response.status_code == 201
    return [kol["id"] for kol in client.get("/kols").json()]


def test_comparison_accepts_up_to_four_kols(client):
    kol_ids = seed_two_kols(client)

    response = client.post("/comparisons", json={"kol_ids": kol_ids})

    assert response.status_code == 200
    assert len(response.json()["items"]) == 2
    assert response.json()["items"][0]["score_summary"] is not None


def test_comparison_rejects_five_kols(client):
    response = client.post(
        "/comparisons",
        json={"kol_ids": [1, 2, 3, 4, 5]},
    )

    assert response.status_code == 422


def test_comparison_reports_missing_kols(client):
    response = client.post("/comparisons", json={"kol_ids": [999]})

    assert response.status_code == 404
    assert "999" in response.json()["detail"]


def test_shortlist_add_remove_and_export(client):
    kol_id = seed_two_kols(client)[0]
    shortlist = client.post(
        "/shortlists",
        json={"name": "UK launch", "target_country": "UK"},
    ).json()
    response = client.post(
        f"/shortlists/{shortlist['id']}/items",
        json={
            "kol_id": kol_id,
            "priority": 1,
            "recommendation": "Strong EV audience",
        },
    )
    assert response.status_code == 201
    shortlist_detail = client.get(f"/shortlists/{shortlist['id']}").json()
    assert shortlist_detail["items"][0]["kol"]["name"] == "Alex EV"
    assert shortlist_detail["items"][0]["kol"]["handle"] == "@alex"

    export = client.get(f"/shortlists/{shortlist['id']}/export")
    assert export.status_code == 200
    assert export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[1][1] == "Alex EV"

    removed = client.delete(
        f"/shortlists/{shortlist['id']}/items/{kol_id}"
    )
    assert removed.status_code == 204
    assert client.get(f"/shortlists/{shortlist['id']}").json()["items"] == []


def test_shortlist_rejects_duplicate_kol(client):
    kol_id = seed_two_kols(client)[0]
    shortlist_id = client.post(
        "/shortlists",
        json={"name": "Launch candidates"},
    ).json()["id"]
    payload = {"kol_id": kol_id}
    assert client.post(
        f"/shortlists/{shortlist_id}/items",
        json=payload,
    ).status_code == 201

    duplicate = client.post(
        f"/shortlists/{shortlist_id}/items",
        json=payload,
    )

    assert duplicate.status_code == 409
